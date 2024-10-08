from io import BytesIO
from itertools import cycle, repeat
from unittest.mock import MagicMock, patch

from django.contrib.auth.models import Group
from django.test import TestCase
from django.utils.html import escape
from django_webtest import WebTest
from model_bakery import baker
from openpyxl import load_workbook

from evap.evaluation.models import Contribution, Course, Evaluation, UserProfile
from evap.evaluation.tests.tools import assert_no_database_modifications
from evap.rewards.models import RewardPointGranting, RewardPointRedemption
from evap.staff.fixtures.excel_files_test_data import (
    create_memory_csv_file,
    create_memory_excel_file,
    valid_user_courses_import_filedata,
    valid_user_courses_import_users,
)
from evap.staff.tools import (
    conditional_escape,
    merge_users,
    remove_user_from_represented_and_ccing_users,
    user_edit_link,
)
from evap.tools import assert_not_none
from tools.enrollment_preprocessor import run_preprocessor


class MergeUsersTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user1 = baker.make(UserProfile, email="test1@institution.example.com")
        cls.user2 = baker.make(UserProfile, email="test2@institution.example.com")
        cls.user3 = baker.make(UserProfile, email="test3@institution.example.com")
        cls.group1 = Group.objects.get(name="Reviewer")
        cls.group2 = Group.objects.get(name="Grade publisher")
        cls.main_user = baker.make(
            UserProfile,
            title="Dr.",
            first_name_given="Main",
            first_name_chosen="",
            last_name="",
            email=None,  # test that merging works when taking the email from other user (UniqueConstraint)
            groups=[cls.group1],
            delegates=[cls.user1, cls.user2],
            represented_users=[cls.user3],
            cc_users=[cls.user1],
            ccing_users=[],
        )
        cls.other_user = baker.make(
            UserProfile,
            title="",
            first_name_given="Other",
            first_name_chosen="other-display-name",
            last_name="User",
            email="other@test.com",
            groups=[cls.group2],
            delegates=[cls.user3],
            represented_users=[cls.user1],
            cc_users=[],
            ccing_users=[cls.user1, cls.user2],
            is_superuser=True,
        )
        cls.course1 = baker.make(Course, responsibles=[cls.main_user])
        cls.course2 = baker.make(Course, responsibles=[cls.main_user])
        cls.course3 = baker.make(Course, responsibles=[cls.other_user])
        cls.evaluation1 = baker.make(
            Evaluation, course=cls.course1, name_de="evaluation1", participants=[cls.main_user, cls.other_user]
        )  # this should make the merge fail
        cls.evaluation2 = baker.make(
            Evaluation, course=cls.course2, name_de="evaluation2", participants=[cls.main_user], voters=[cls.main_user]
        )
        cls.evaluation3 = baker.make(
            Evaluation,
            course=cls.course3,
            name_de="evaluation3",
            participants=[cls.other_user],
            voters=[cls.other_user],
        )
        cls.contribution1 = baker.make(Contribution, contributor=cls.main_user, evaluation=cls.evaluation1)
        cls.contribution2 = baker.make(
            Contribution, contributor=cls.other_user, evaluation=cls.evaluation1
        )  # this should make the merge fail
        cls.contribution3 = baker.make(Contribution, contributor=cls.other_user, evaluation=cls.evaluation2)
        cls.rewardpointgranting_main = baker.make(RewardPointGranting, user_profile=cls.main_user)
        cls.rewardpointgranting_other = baker.make(RewardPointGranting, user_profile=cls.other_user)
        cls.rewardpointredemption_main = baker.make(RewardPointRedemption, user_profile=cls.main_user)
        cls.rewardpointredemption_other = baker.make(RewardPointRedemption, user_profile=cls.other_user)

    def test_merge_handles_all_attributes(self):
        user1 = baker.make(UserProfile)
        user2 = baker.make(UserProfile)

        all_attrs = [field.name for field in UserProfile._meta.get_fields(include_hidden=True)]

        # these are relations to intermediate models generated by django for m2m relations.
        # we can safely ignore these since the "normal" fields of the m2m relations are present as well.
        all_attrs = [attr for attr in all_attrs if not attr.startswith("UserProfile_")]

        # equally named fields are not supported, sorry
        self.assertEqual(len(all_attrs), len(set(all_attrs)))

        # some attributes we don't care about when merging
        ignored_attrs = {
            "id",  # nothing to merge here
            "password",  # not used in production
            "last_login",  # something to really not care about
            "user_permissions",  # we don't use permissions
            "logentry",  # wtf
            "login_key",  # we decided to discard other_user's login key
            "login_key_valid_until",  # not worth dealing with
            "language",  # Not worth dealing with
            "Evaluation_voters+",  # some more intermediate models, for an explanation see above
            "Evaluation_participants+",  # intermediate model
            "startpage",  # not worth dealing with
        }
        expected_attrs = set(all_attrs) - ignored_attrs

        # actual merge happens here
        merged_user, errors, warnings = merge_users(user1, user2)
        self.assertEqual(errors, [])
        self.assertEqual(warnings, [])
        handled_attrs = set(merged_user.keys())

        # attributes that are handled in the merge method but that are not present in the merged_user dict
        # add attributes here only if you're actually dealing with them in merge_users().
        additional_handled_attrs = {
            "grades_last_modified_user+",
            "Course_responsibles+",
        }

        actual_attrs = handled_attrs | additional_handled_attrs

        self.assertEqual(expected_attrs, actual_attrs)

    def test_merge_users_does_not_change_data_on_fail(self):
        with assert_no_database_modifications():
            __, errors, warnings = merge_users(self.main_user, self.other_user)  # merge should fail

        self.assertCountEqual(errors, ["contributions", "evaluations_participating_in"])
        self.assertCountEqual(warnings, ["rewards"])

    def test_merge_users_changes_data_on_success(self):
        # Fix data so that the merge will not fail as in test_merge_users_does_not_change_data_on_fail
        self.evaluation1.participants.set([self.main_user])
        self.contribution2.delete()

        __, errors, warnings = merge_users(self.main_user, self.other_user)  # merge should succeed
        self.assertEqual(errors, [])
        self.assertEqual(warnings, ["rewards"])  # rewards warning is still there

        self.main_user.refresh_from_db()

        self.assertEqual(self.main_user.title, "Dr.")
        self.assertEqual(self.main_user.first_name_given, "Main")
        self.assertEqual(self.main_user.first_name_chosen, "other-display-name")
        self.assertEqual(self.main_user.last_name, "User")
        self.assertEqual(self.main_user.email, "other@test.com")
        self.assertTrue(self.main_user.is_superuser)
        self.assertEqual(set(self.main_user.groups.all()), {self.group1, self.group2})
        self.assertEqual(set(self.main_user.delegates.all()), {self.user1, self.user2, self.user3})
        self.assertEqual(set(self.main_user.represented_users.all()), {self.user1, self.user3})
        self.assertEqual(set(self.main_user.cc_users.all()), {self.user1})
        self.assertEqual(set(self.main_user.ccing_users.all()), {self.user1, self.user2})
        self.assertTrue(RewardPointGranting.objects.filter(user_profile=self.main_user).exists())
        self.assertTrue(RewardPointRedemption.objects.filter(user_profile=self.main_user).exists())

        self.assertEqual(set(self.course1.responsibles.all()), {self.main_user})
        self.assertEqual(set(self.course2.responsibles.all()), {self.main_user})
        self.assertEqual(set(self.course2.responsibles.all()), {self.main_user})
        self.assertEqual(set(self.evaluation1.participants.all()), {self.main_user})
        self.assertEqual(set(self.evaluation2.participants.all()), {self.main_user})
        self.assertEqual(set(self.evaluation2.voters.all()), {self.main_user})
        self.assertEqual(set(self.evaluation3.participants.all()), {self.main_user})
        self.assertEqual(set(self.evaluation3.voters.all()), {self.main_user})

        self.assertFalse(UserProfile.objects.filter(email="other_user@institution.example.com").exists())
        self.assertFalse(RewardPointGranting.objects.filter(user_profile__email=self.other_user.email).exists())
        self.assertFalse(RewardPointRedemption.objects.filter(user_profile__email=self.other_user.email).exists())


class RemoveUserFromRepresentedAndCCingUsersTest(TestCase):
    def test_remove_user_from_represented_and_ccing_users(self):
        delete_user = baker.make(UserProfile)
        delete_user2 = baker.make(UserProfile)
        user1 = baker.make(UserProfile, delegates=[delete_user, delete_user2], cc_users=[delete_user])
        user2 = baker.make(UserProfile, delegates=[delete_user], cc_users=[delete_user, delete_user2])

        messages = remove_user_from_represented_and_ccing_users(delete_user)
        self.assertEqual([set(user1.delegates.all()), set(user1.cc_users.all())], [{delete_user2}, set()])
        self.assertEqual([set(user2.delegates.all()), set(user2.cc_users.all())], [set(), {delete_user2}])
        self.assertEqual(len(messages), 4)

        messages2 = remove_user_from_represented_and_ccing_users(delete_user2)
        self.assertEqual([set(user1.delegates.all()), set(user1.cc_users.all())], [set(), set()])
        self.assertEqual([set(user2.delegates.all()), set(user2.cc_users.all())], [set(), set()])
        self.assertEqual(len(messages2), 2)

    def test_do_not_remove_from_ignored_users(self):
        delete_user = baker.make(UserProfile)
        user1 = baker.make(UserProfile, delegates=[delete_user], cc_users=[delete_user])
        user2 = baker.make(UserProfile, delegates=[delete_user], cc_users=[delete_user])

        messages = remove_user_from_represented_and_ccing_users(delete_user, [user2])
        self.assertEqual([set(user1.delegates.all()), set(user1.cc_users.all())], [set(), set()])
        self.assertEqual([set(user2.delegates.all()), set(user2.cc_users.all())], [{delete_user}, {delete_user}])
        self.assertEqual(len(messages), 2)

    def test_do_nothing_if_test_run(self):
        delete_user = baker.make(UserProfile)
        user1 = baker.make(UserProfile, delegates=[delete_user], cc_users=[delete_user])
        user2 = baker.make(UserProfile, delegates=[delete_user], cc_users=[delete_user])

        messages = remove_user_from_represented_and_ccing_users(delete_user, test_run=True)
        self.assertEqual([set(user1.delegates.all()), set(user1.cc_users.all())], [{delete_user}, {delete_user}])
        self.assertEqual([set(user2.delegates.all()), set(user2.cc_users.all())], [{delete_user}, {delete_user}])
        self.assertEqual(len(messages), 4)


class UserEditLinkTest(TestCase):
    def test_user_edit_link(self):
        user = baker.make(UserProfile)
        self.assertIn(f"/staff/user/{user.id}/edit", user_edit_link(user.id))


class ConditionalEscapeTest(TestCase):
    def test_conditional_escape(self):
        self.assertEqual(conditional_escape("<script>"), "&lt;script&gt;")
        self.assertEqual(conditional_escape(escape("<script>")), "&lt;script&gt;")
        self.assertEqual(conditional_escape("safe"), "safe")


@patch("tools.enrollment_preprocessor._stdout")
class EnrollmentPreprocessorTest(WebTest):
    @classmethod
    def setUpTestData(cls) -> None:
        cls.imported_data = valid_user_courses_import_filedata
        cls.csv = create_memory_csv_file(
            [["Title", "Last name", "First name", "Email"]] + valid_user_courses_import_users
        )

    @patch("builtins.input", side_effect=cycle(("i", "e", "invalid")))
    def test_xlsx_data_stripped(self, input_patch: MagicMock, _stdout_patch):
        self.imported_data["MA Belegungen"][1][1] = " Accepted  "
        self.imported_data["MA Belegungen"][1][8] = "   conflicts  "
        self.imported_data["BA Belegungen"][1][2] = "   are    "
        self.imported_data["BA Belegungen"][1][11] = "   stripped.   "
        modified = run_preprocessor(BytesIO(create_memory_excel_file(self.imported_data)), self.csv)
        self.assertIsNotNone(modified)
        self.assertEqual(input_patch.call_count, 4)
        workbook = load_workbook(assert_not_none(modified), read_only=True)
        self.assertEqual(workbook["MA Belegungen"]["B2"].value, "Accepted")  # stripped conflict used
        self.assertEqual(workbook["MA Belegungen"]["I2"].value, None)  # existing data kept
        self.assertEqual(workbook["BA Belegungen"]["C2"].value, "are")  # stripped conflict used
        self.assertEqual(workbook["BA Belegungen"]["L2"].value, "stripped.")  # different email is no conflict

    @patch("builtins.input", side_effect=repeat("i"))
    def test_empty_email_ignored(self, input_patch: MagicMock, _stdout_patch):
        self.imported_data["MA Belegungen"][1][1] = " Add  "
        self.imported_data["MA Belegungen"][1][8] = "   some  "
        self.imported_data["BA Belegungen"][1][2] = "   conflicts    "
        self.imported_data["MA Belegungen"][1][3] = " "
        self.imported_data["MA Belegungen"][1][11] = "   "
        self.imported_data["BA Belegungen"][1][3] = " \t    "
        self.imported_data["BA Belegungen"][1][11] = "  \n   "
        res = run_preprocessor(BytesIO(create_memory_excel_file(self.imported_data)), self.csv)
        self.assertIsNone(res)
        input_patch.assert_not_called()

    @patch("builtins.input", side_effect=repeat("i"))
    def test_deduplication(self, input_patch: MagicMock, _stdout_patch):
        self.imported_data["MA Belegungen"][1][1] = "Some conflicts"
        self.imported_data["MA Belegungen"][1][8] = "in all"
        self.imported_data["BA Belegungen"][1][2] = "fields"
        # copy data and pad with spaces
        self.imported_data["MA Belegungen"].append([f" {data} " for data in self.imported_data["MA Belegungen"][1]])

        res = run_preprocessor(BytesIO(create_memory_excel_file(self.imported_data)), self.csv)
        self.assertIsNone(res)
        self.assertEqual(input_patch.call_count, 3)  # conflicts are deduplicated.

    @patch("builtins.input", side_effect=cycle(("i", "e", "e", "invalid")))
    def test_changes_applied_globally(self, input_patch: MagicMock, _stdout_patch):
        self.imported_data["MA Belegungen"][1][1] = "some conflicts"
        self.imported_data["MA Belegungen"][1][8] = "in all"
        self.imported_data["BA Belegungen"][1][2] = "fields"
        # copy data and pad with spaces and add conflict
        self.imported_data["MA Belegungen"].append([f" {data} " for data in self.imported_data["MA Belegungen"][1]])
        self.imported_data["BA Belegungen"].append([f" {data} " for data in self.imported_data["BA Belegungen"][1]])
        self.imported_data["MA Belegungen"][2][1] += "modified"
        self.imported_data["MA Belegungen"][2][8] += "modified"
        self.imported_data["BA Belegungen"][2][2] += "modified"
        self.imported_data["MA Belegungen"].append([f" {data} " for data in self.imported_data["MA Belegungen"][2]])
        self.imported_data["BA Belegungen"].append([f" {data} " for data in self.imported_data["BA Belegungen"][2]])
        modified = run_preprocessor(BytesIO(create_memory_excel_file(self.imported_data)), self.csv)
        self.assertIsNotNone(modified)
        self.assertEqual(input_patch.call_count, 7)
        workbook = load_workbook(assert_not_none(modified), read_only=True)
        self.assertEqual(workbook["MA Belegungen"]["B2"].value, "some conflicts")
        self.assertEqual(workbook["MA Belegungen"]["B3"].value, "some conflicts")
        self.assertEqual(workbook["MA Belegungen"]["I2"].value, "in all modified")
        self.assertEqual(workbook["MA Belegungen"]["I3"].value, "in all modified")
        self.assertEqual(workbook["BA Belegungen"]["C2"].value, "Lucilia")
        self.assertEqual(workbook["BA Belegungen"]["C3"].value, "Lucilia")
